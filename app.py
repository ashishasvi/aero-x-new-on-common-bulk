from flask import Flask, request, render_template, send_from_directory
import openpyxl
from datetime import date
import pandas as pd
import os

app = Flask(__name__)

# Excel new bulk formatter change
def reorder_csv_to_bulk_format(csv_file_path, output_excel_path):
    # Load the CSV file
    csv_df = pd.read_csv(csv_file_path)

    # Desired order of columns (final columns to keep)
    final_order = [
        'Name',
        'inscor__Product__r.Name',
        'inscor__Condition_Code__r.Name',
        'inscor__Quantity_Available__c',
        'inscor__UOM__c',  # This column will be added with a default value
        'SSP_Updated__c',
        'inscor__Keyword__c'
    ]

    # Retain only the columns that are in the final_order
    csv_df = csv_df[[col for col in final_order if col in csv_df.columns]]

    # Add the missing 'inscor__UOM__c' column with default value "EA"
    if 'inscor__UOM__c' not in csv_df.columns:
        csv_df.loc[:, 'inscor__UOM__c'] = "EA"  # Using .loc to avoid SettingWithCopyWarning

    # Reorder columns as per the final order, including the new 'inscor__UOM__c' column
    csv_reordered = csv_df.reindex(columns=final_order)

    # Ensure the output path has the correct file extension
    if not output_excel_path.lower().endswith('.xlsx'):
        output_excel_path += '.xlsx'

    # Save the reordered CSV as an Excel file
    csv_reordered.to_excel(output_excel_path, index=False)

# Excel formatter change end

def formatmonth(date_obj):
    monthNames = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    month = monthNames[date_obj.month - 1]
    return month

def process_excel(input_file_path, output_file_path):
    # Load the input workbook
    workbook = openpyxl.load_workbook(input_file_path)

    # Create a new worksheet with name "template"
    template = workbook.create_sheet("template")

    # Get the source worksheet "bulk"
    selected_sheet = workbook.worksheets[0]

    # Copy data from selected_sheet to template
    for row in selected_sheet.iter_rows(min_row=1, values_only=True):
        template.append((None, None, None, row[1], None, None, None, row[6], None, None, None, None, None, None, None, row[2], row[3], row[4], None, row[5]))

    # REPLACE RAI TD ST TO AR CONDITION
    for cell in template["P"]:
        if cell.value in {"RAI", "TD", "ST"}:
            cell.value = "AR"

    # REPLACE SPECIAL CHARACTER IN PARTS NUMBER
    for cell in template["D"]:
        cell.value = str(cell.value).replace("/", "").replace(" ", "")

    # REPLACE BLANK DESCRIPTION WITH NO DESCRIPTION
    for cell in template["H"]:
        if cell.value == " ":
            cell.value = "NO DESCRIPTION"

    # PUT EA USD AND DATES
    today = date.today()
    next_year = today.replace(year=today.year + 1)
    datenow = f"{today.day}-{formatmonth(today)}-{today.year}"
    datenextyear = f"{next_year.day}-{formatmonth(next_year)}-{next_year.year}"

    data_to_write = [["USD", datenow, datenextyear] for _ in range(1, template.max_row)]
    for row_idx, row in enumerate(data_to_write, start=2):
        for col_idx, value in enumerate(row, start=21):  # Column U starts at index 21
            template.cell(row=row_idx, column=col_idx, value=value)

    # Set date format for columns V and W
    for col in ["V", "W"]:
        for cell in template[col]:
            cell.number_format = "dd-mmm-yyyy"

    for i, letter in enumerate("ABCDEFGHIJKLMNOPQRSTUVWX", start=1):
        template[f"{letter}1"] = letter

    workbook.remove(workbook.worksheets[0])
    workbook.save(output_file_path)
    workbook.close()

    # Read the Excel sheet into a pandas dataframe
    df = pd.read_excel(output_file_path)

    # Find duplicates based on columns 'D' and 'P', and sum up the quantity values
    duplicates = df[df.duplicated(subset=['D', 'P'], keep=False)].groupby(['D', 'P']).agg({'Q': 'sum', 'A': 'first', 'B': 'first', 'C': 'first', 'E': 'first', 'F': 'first', 'G': 'first', 'H': 'first', 'I': 'first', 'J': 'first', 'K': 'first', 'L': 'first', 'M': 'first', 'N': 'first', 'O': 'first', 'R': 'first', 'S': 'first', 'T': 'first', 'U': 'first', 'V': 'first', 'W': 'first'}).reset_index()

    # Drop duplicates based on columns 'D' and 'P'
    df = df.drop_duplicates(subset=['D', 'P'], keep=False)

    # Concatenate the original dataframe and the summed duplicates
    df = pd.concat([df, duplicates], ignore_index=True)

    # Save the modified dataframe back to Excel, overwriting the original file
    df.to_excel(output_file_path, index=False)

    # Reopen the file after processing
    workbook = openpyxl.load_workbook(output_file_path)
    template_sheet = workbook["Sheet1"]

    # Clear the first row and set specific values
    for cell in template_sheet["1"]:
        cell.value = None

    template_sheet["A1"] = "4.00"
    template_sheet["F1"] = "7RVW9"
    template_sheet["H1"] = "cam@setnaio.com"

    workbook.save(output_file_path)

    # Update column M with specific emails and create text files
    email_updates = [
        ("paul@setnaio.com", "29098"),
        ("Stefan.Shemaitis@setnaio.com", "29099"),
        ("Stefan.Shemaitis@setnaio.com", "29100"),
        ("Stefan.Shemaitis@setnaio.com", "29663"),
        ("paul@setnaio.com", "29665"),
        ("aybars@setnaio.com", "30297"),
        ("Scott.Loza@setnaio.com", "29664"),
        ("paul@setnaio.com", ""),
    ]

    txt_files = []
    for email_value, g_value in email_updates:
        template_sheet["G1"] = g_value
        if g_value == "":
          for row in range(1, template_sheet.max_row + 1):
            template_sheet[f"T{row}"] = ""  # Sets every cell in column T to an empty string


        for row in template_sheet.iter_rows(min_row=2, min_col=13, max_col=13):
            for cell in row:
                
                    cell.value = email_value

        txt_file_path = os.path.join("downloads", f"SetnaiO_7RVW9_{g_value}.txt")
        txt_files.append(txt_file_path)

        with open(txt_file_path, "w", encoding="utf-8") as txt_file:
            for row in template_sheet.iter_rows(values_only=True):
                row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
                txt_file.write(row_str + "\n")

    # Create a special text file for NS and NE conditions
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    # Copy header row to new sheet
    new_sheet.append([cell.value for cell in template_sheet[1]])

    allowed_values = ["NS", "NE"]
    for row in template_sheet.iter_rows(min_row=2, min_col=1, max_col=24):
        cell_value = row[15].value  # Value in column P
        if cell_value in allowed_values:
            new_sheet.append([cell.value for cell in row])

    new_sheet["G1"] = "29664"
    for row in new_sheet.iter_rows(min_row=2, min_col=13, max_col=13):
            for cell in row:
                
                    cell.value = "Scott.Loza@setnaio.com"
    final_txt_file = os.path.join("downloads", "SetnaiO_7RVW9_29664.txt")
    txt_files.append(final_txt_file)

    with open(final_txt_file, "w", encoding="utf-8") as txt_file:
        for row in new_sheet.iter_rows(values_only=True):
            row_str = "\t".join(str(cell) if cell is not None else "" for cell in row)
            txt_file.write(row_str + "\n")

    new_workbook.save(output_file_path)

    return txt_files

@app.route('/download/<path:filename>')
def download_file(filename):
    download_folder = "downloads"  # The folder where your files are located
    file_path = os.path.join(download_folder, filename)
    return send_from_directory(download_folder, filename, as_attachment=True)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('index.html', message='No file selected')

        file = request.files['file']
        if file.filename == '':
            return render_template('index.html', message='No file selected')

        try:
            # For transform bulk
            csv_file_path = os.path.join("uploads", file.filename)
            output_excel_path = os.path.join("uploads", "transform", f"{os.path.splitext(file.filename)[0]}.xlsx")
            
            # For process
            output_file_path = os.path.join("downloads", f"{os.path.splitext(file.filename)[0]}.xlsx")
            file.save(csv_file_path)

            reorder_csv_to_bulk_format(csv_file_path, output_excel_path)
            txt_files = process_excel(output_excel_path, output_file_path)

            if not txt_files:
                return render_template('index.html', message='No files generated.')

            return render_template('index.html', message='File processed successfully', download_file=txt_files)

        except Exception as e:
            return render_template('index.html', message=f'Error: {e}')

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)
